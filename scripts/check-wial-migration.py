#!/usr/bin/env python3
"""Report wial.org → chapterstack migration status.

Usage:
    python3 scripts/check-wial-migration.py            # downloads the crawler sheet
    python3 scripts/check-wial-migration.py --sheet local.xlsx
    npm run check:migration

Cross-references three sources of truth and prints a markdown status report:

1. The website crawler spreadsheet (Google Sheet, downloaded as xlsx so the
   color groupings survive; colors mark which legacy pages merge together).
   Every crawled wial.org URL is classified: redirected via aliasMap, live as
   a canonical page, tracked by a GitHub issue, or UNTRACKED.
2. `src/lib/routing.ts` — aliasMap (legacy path redirects) and canonicalMap.
3. `src/content/pages.json` — every wial.org URL still embedded in content.
   URLs inside a page's `seo` block are provenance (fine); anywhere else is
   an outbound link a visitor can click, which must be tracked by an issue
   or explained in `seo.sourceNotes`.

If `gh` is installed, referenced issues are annotated OPEN/CLOSED. Exits 1
when anything is untracked, so the check can gate CI or a pre-release list.

Maintaining: when new migration issues are filed (or paths ship), update
ISSUE_MAP / OUTBOUND_HINTS below and the table in rules/wial-migration.md.
"""
import json
import re
import shutil
import subprocess
import sys
import tempfile
import urllib.request
from pathlib import Path
from urllib.parse import urlparse

REPO_ROOT = Path(__file__).resolve().parent.parent
SHEET_ID = "18feQ3-v9ZropFo19fgT7WVbJ-dBM6EuE"
SHEET_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
GITHUB_REPO = "2026-ASU-WiCS-Opportunity-Hack/03-nanpossible"

# Legacy path (no leading/trailing slash, lowercase) -> GitHub issue tracking it.
ISSUE_MAP = {
    "": 9,  # homepage exists; content refresh tracked
    "certification": 100,  # page live; sub-page merge tracked
    "certification/foundations": 100,
    "certification/in-house-programs": 100,
    "certification/calc-courses": 100,
    "programs": 100,
    "become-a-coach": 100,
    "become-a-partner": 117,
    "become-an-affiliate": 118,
    "wials-team": 119,
    "wial-talk": 120,
    "action-learning/library": 120,
    "wial-endorsed-products": 120,
    "wp-content/uploads/2026/02/power_of_action_learning.pdf": 120,
    "wp-content/uploads/2026/02/wial-action-learning-brochure-1.pdf": 120,
    "category/wial-blog": 121,
    "privacy-policy": 122,
}

# Substring of an outbound wial.org URL in content -> issue tracking its removal.
OUTBOUND_HINTS = {
    "become-a-partner": 117,
    "wial-better-world-fund-donation": 123,
    "wial-better-world-fund-application": 123,
    "/projects/": 123,
}


def download_sheet(dest: Path) -> None:
    req = urllib.request.Request(SHEET_EXPORT_URL, headers={"User-Agent": "wial-migration-check"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            dest.write_bytes(resp.read())
    except Exception as err:  # noqa: BLE001
        sys.exit(
            f"Could not download the crawler sheet ({err}).\n"
            "Download it manually (File > Download > .xlsx) and pass --sheet <path>."
        )


def load_sheet_rows(xlsx_path: Path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip3 install openpyxl")
    ws = openpyxl.load_workbook(xlsx_path)[
        "crawl_wial.org"
    ]
    header = {cell.value: cell.column for cell in ws[1] if cell.value}
    rows = []
    for row in ws.iter_rows(min_row=2):
        def val(name):
            col = header.get(name)
            v = row[col - 1].value if col else None
            return str(v).strip() if v is not None else ""

        url = val("Full URL")
        if not url:
            continue
        rows.append(
            {
                "row": row[0].row,
                "url": url,
                "title": val("Title") or val("Page"),
                "note": val("Steps to do on Chapterstack site"),
            }
        )
    return rows


def parse_routing():
    text = (REPO_ROOT / "src/lib/routing.ts").read_text()

    def entries(map_name):
        block = re.search(rf"const {map_name} = new Map[^(]*\(\[(.*?)\]\);", text, re.S)
        if not block:
            sys.exit(f"Could not find {map_name} in src/lib/routing.ts")
        return re.findall(r'\["([^"]*)",\s*"([^"]+)"\]', block.group(1))

    alias = dict(entries("aliasMap"))
    canonical = {k for k, _ in entries("canonicalMap")}
    return alias, canonical


def issue_states():
    if not shutil.which("gh"):
        return {}
    try:
        out = subprocess.run(
            ["gh", "issue", "list", "--repo", GITHUB_REPO, "--state", "all",
             "--limit", "300", "--json", "number,state"],
            capture_output=True, text=True, check=True, timeout=30,
        ).stdout
        return {item["number"]: item["state"] for item in json.loads(out)}
    except Exception:  # noqa: BLE001
        return {}


def issue_label(num, states):
    state = states.get(num)
    return f"#{num} ({state})" if state else f"#{num}"


def classify_path(path, alias, canonical, states):
    tracked = ISSUE_MAP.get(path)
    if path in alias:
        return f"redirects to `{alias[path]}`", True
    if path in canonical:
        target = f"/{path}" if path else "/"
        if tracked:
            return f"live at `{target}`, content tracked in {issue_label(tracked, states)}", True
        return f"live at `{target}`", True
    if tracked:
        return f"tracked in {issue_label(tracked, states)}", True
    return "UNTRACKED", False


def scan_outbound_links():
    """Yield (location, url, in_seo) for every wial.org URL in pages.json."""
    pages = json.loads((REPO_ROOT / "src/content/pages.json").read_text())
    found = []

    def walk(node, crumbs, in_seo):
        if isinstance(node, dict):
            for key, value in node.items():
                walk(value, crumbs + [key], in_seo or key == "seo")
        elif isinstance(node, list):
            for i, value in enumerate(node):
                walk(value, crumbs + [str(i)], in_seo)
        elif isinstance(node, str):
            for url in re.findall(r"https?://(?:www\.)?wial\.org[^\s\"')]*", node):
                found.append((".".join(crumbs), url, in_seo))

    for page in pages:
        walk(page, [f"pages.json:{page.get('slug', '?')}"], False)
    return found


def main():
    args = sys.argv[1:]
    if "--sheet" in args:
        xlsx = Path(args[args.index("--sheet") + 1])
    else:
        xlsx = Path(tempfile.mkstemp(suffix=".xlsx")[1])
        print(f"Downloading crawler sheet {SHEET_ID}...", file=sys.stderr)
        download_sheet(xlsx)

    alias, canonical = parse_routing()
    states = issue_states()
    rows = load_sheet_rows(xlsx)

    print("# wial.org migration status\n")
    print("## Crawler sheet coverage\n")
    print("| Sheet row | Legacy path | Status |")
    print("| --- | --- | --- |")
    untracked = 0
    done = 0
    for entry in rows:
        path = urlparse(entry["url"]).path.strip("/").lower()
        status, ok = classify_path(path, alias, canonical, states)
        if not ok:
            untracked += 1
            status = "**UNTRACKED** (no redirect, no page, no issue)"
        elif status.startswith(("redirects", "live at `/`")) or (
            status.startswith("live") and "tracked" not in status
        ):
            done += 1
        print(f"| {entry['row']} | `/{path}` | {status} |")

    print("\n## Outbound wial.org links still in content\n")
    outbound = [(loc, url) for loc, url, in_seo in scan_outbound_links() if not in_seo]
    unexplained = 0
    if not outbound:
        print("None. Content stands on its own.")
    else:
        print("| Location | URL | Status |")
        print("| --- | --- | --- |")
        for loc, url in outbound:
            issue = next((n for frag, n in OUTBOUND_HINTS.items() if frag in url), None)
            if issue:
                status = f"tracked in {issue_label(issue, states)}"
            else:
                unexplained += 1
                status = "**UNTRACKED** (add to OUTBOUND_HINTS or record in sourceNotes)"
            print(f"| {loc} | {url} | {status} |")

    total = len(rows)
    print("\n## Summary\n")
    print(f"- Sheet rows fully migrated (live or redirected, nothing pending): {done}/{total}")
    print(f"- Sheet rows tracked by an open/closed issue: {total - done - untracked}")
    print(f"- Sheet rows untracked: {untracked}")
    print(f"- Outbound content links: {len(outbound)} ({unexplained} untracked)")
    if not states:
        print("- (install/auth `gh` to see issue open/closed states)")

    sys.exit(1 if (untracked or unexplained) else 0)


if __name__ == "__main__":
    main()
